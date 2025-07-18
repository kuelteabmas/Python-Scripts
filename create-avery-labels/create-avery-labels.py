from docxtpl import DocxTemplate
from openpyxl import load_workbook
# from openpyxl.cell.text import InlineFont, Font
# from openpyxl.cell.rich_text import TextBlock, CellRichText



def iterating_column(path, sheet_name, col):

    # # Font size = 10 
    # # fontForOver23Chars = InlineFont(sz="10.0")
    # fontForOver23Chars = InlineFont(sz=10)
    

    # Import template document
    template = DocxTemplate('avery-template-5167-font-size-8.docx')

    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return
        
    sheet = workbook[sheet_name]
    for cell in sheet[col]:

        # replace chars in fetched cell result
        cellData = cell.value

        # if len(cellData) > 23:
        #     cellData = CellRichText(
        #         # cellData,
        #         TextBlock(fontForOver23Chars, cellData)
        #     )

        s = str(cellData)
        s.translate({ord('\''):None}) # remove single quotes from cell data object
        print("Name: "+ s)

        # Declare template variables
        context = {
            'name': s
        }


        # # if char size of label name/cell data is < 23, print 'type' at the end of the label name 
        # if len(cellData) < 23:
        #     context = {
        #         'name': s
        #     }

        # empty_str = " "
        # # if char size of label name/cell data is > 23, don't print 'type' at the end of the label name 
        # if len(cellData) > 23:
        #     context = {
        #         'name': s,
        #         'type': empty_str
        #     }


        # Copy data from variables to template variables
        template.render(context)

        # Write changes to new document
        fileName = "%s Type.docx"%(s)
        template.save(fileName)

if __name__ == "__main__":
    iterating_column("output.xlsx", sheet_name="Sheet1", col="A")
